import logging
import openpyxl as xl
import datetime as dt
import calendar
from difflib import SequenceMatcher

logging.basicConfig(filename='mylog.log', level=logging.DEBUG)
file_to_open = 'expedia_report_monthly_january_2018.xlsx'
LABEL_INFLEXIBILITY = 0.8
SHEET_NAME_INFLEXIBILITY = 0.8
TEST_SHEET = [
    'Summary Rolling MoM',
    ('Calls Offered', 'int'),
    ('Abandon after 30s', 'pct'),
    ('FCR', 'pct'),
    ('DSAT', 'pct'),
    ('CSAT', 'pct')
]
VOC_SHEET = [
    'VOC Rolling MoM',
    ('Base Size', 'int'),
    ('Promoters (Recommend Score 9 to 10)', 'int, "good" if val > 200 else "bad"'),
    ('Passives (Recommend Score 7 to 8)', 'int, "good" if val > 100 else "bad"'),
    ('Detractors (recommend Score 0 to 6)', 'int, "good" if val < 100 else "bad"'),
    ('Overall NPS %', 'pct', 1),
    ('Sat with Agent %', 'pct', 1),
    ('DSat with Agent %', 'pct', 1)
]
COMPARISON_OPERATORS = [
    '==',
    '>',
    '<',
    '>=',
    '<=',
    '!='
]


# replace every item in stop_chars within s with rep
def replaceAll(s, stop_chars, rep):
    for ch in stop_chars:
        s = s.replace(ch, rep)
    return s


# Check whether cont contains any of elems
def contains_any(cont, elems):
    ret = False
    for elem in elems:
        ret = ret or elem in cont
    return ret


# Get the excel column letter from its number
def get_column_letter(col_num):
    # alphabet for indexing
    alph = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    # get the first letter
    ltr = alph[col_num % 26 - 1]
    col_num -= 27
    # iteratively get other letters if necessary
    mult = 26
    while col_num > 0:
        ltr = alph[(col_num // mult) % 26] + ltr
        mult *= 26
        col_num -= mult
    return ltr


# get the month/year from the file name
def get_date(file_name):
    month = None
    year = None
    # checking lower against lower for ease
    fn_lower = file_name.lower()
    for i in range(1, 13):
        # check every full month
        mo = calendar.month_name[i]
        if mo.lower() in fn_lower:
            month = i
            # break to avoid else block
            break
    else:
        # if we don't find a full month, try the 3-letter abbreviation
        for i in range(1, 13):
            # check every month
            mo = calendar.month_name[i]
            if mo.lower()[:3] in fn_lower:
                month = i
                # break to avoid else block
                break
        else:
            # log that we couldn't find a month
            logging.exception(f'No month in file name: {file_name}')
    # check every 4-letter substring to find one that can be an int
    for i in range(len(fn_lower) - 4):
        try:
             yr = int(fn_lower[i:i+4])
        except ValueError:
            continue
        else:
            year = yr
    # if there was no 4-letter substring that could cast to int, log that we couldn't find a year
    if not year:
        logging.exception(f'No year in file name : {file_name}')
    return month, year


# check that the file exists
def file_exists(file_name):
    try:
        with open(file_name):
            pass
    except FileNotFoundError:
        logging.exception(f'File {file_name} does not exist')
    else:
        return True


# get the sheet from wb named name
def get_sheet(wb, name):
    # if the name is verbatim in the workbook, get the sheet
    if name in wb.sheetnames:
        return wb[name]
    # otherwise, check all the sheets that actually exist
    for real_name in wb.sheetnames:
        # check to find the closest match
        if SequenceMatcher(a=real_name, b=name).ratio() > SHEET_NAME_INFLEXIBILITY:
            return wb[real_name]
    logging.error(f'Unable to find sheet with name {name}')


# essentially just a container for different case data
class DataPoint:

    def __init__(self, label, form, offset=0):
        self.label = label
        self.form = form
        self.offset = offset


# main logging class for parsing the data
class DataLogger:

    def __init__(self, file_name, cases):
        self.file_name = file_name
        self.cases = cases
        exists = file_exists(file_name)
        if exists:
            self.wb = xl.load_workbook(file_name, data_only=True)

    # Get the addresses of the different labels
    @staticmethod
    def find_labels(sheet, labels):
        # get just the str(label)
        labels = [l.label for l in labels]
        # addresses[label] = [column, row]
        addresses = {}
        header_rows = []
        # find all exact match labels by iterating through every row
        for row in sheet.iter_rows():
            # in this list form we can use the in keyword
            l_row = [cell.value for cell in row]
            for l in labels:
                if l in l_row:
                    # add the row to the headers for later checking
                    if row not in header_rows: header_rows.append(row)
                    cell = row[l_row.index(l)]
                    addresses[l] = [cell.column, cell.row]
        # labels that were not found in the first pass - didn't exact match
        labels_to_find = []
        for l in labels:
            if l not in addresses:
                labels_to_find.append(l)
        # get all the columns which had a label in them for double checking
        addrMin = min([x[0] for x in addresses.values()]) - 1
        addrMax = max([x[0] for x in addresses.values()])
        for h in list(sheet.iter_cols())[addrMin:addrMax]:
            header_rows.append(h)
        # double check all the useful rows
        for row in header_rows:
            for cell in row:
                # if we don't have an exact match and the cell isn't a string, we don't need it
                if type(cell.value) != str:
                    continue
                for l in labels_to_find:
                    # check whether the cell value is a close match to the label
                    if SequenceMatcher(a=cell.value, b=l).ratio() >= LABEL_INFLEXIBILITY or l in cell.value:
                        addresses[l] = [cell.column, cell.row]
                        labels_to_find.remove(l)
        # log any labels that were not found
        for l in labels_to_find:
            logging.exception(f'Unable to find label: {l} in sheet {sheet.title}')
        return addresses

    # Get the address of the date header
    def find_date(self, sheet, month, year, direction):
        if direction == 'h':
            data = sheet.iter_cols()
        else:
            data = sheet.iter_rows()
        for lis in data:
            for cell in lis:
                val = cell.value
                if type(val) is dt.date or type(val) is dt.datetime:
                    if val.month == month and val.year == year:
                        return cell.column, cell.row
        logging.exception(f'Unable to find label for {month}/{year}')

    # Collect and return all the data
    def collect_data(self, index, month, year):
        case = self.cases[index]
        sheet, labels = case[0], case[1:]
        sheet = get_sheet(self.wb, sheet)
        addresses = self.find_labels(sheet, labels)
        tot_col, tot_row = sum([x[0] for x in addresses.values()]), sum([x[1] for x in addresses.values()])
        direction = 'h' if tot_col > tot_row else 'v'
        date_addr = self.find_date(sheet, month, year, direction)
        data = []
        if direction == 'h':
            for lb, addr in addresses.items():
                offset = sum(x.offset for x in labels if x.label == lb)
                data.append([lb, sheet[get_column_letter(addr[0] + offset) + str(date_addr[1])].value])
        else:
            for lb, addr in addresses.items():
                offset = sum(x.offset for x in labels if x.label == lb)
                data.append([lb, sheet[get_column_letter(date_addr[0]) + str(addr[1] + offset)].value])
        return data

    def log_data(self):
        index = 0

        def sort_func(it):
            return [l.label for l in self.cases[index][1:]].index(it[0])

        def map_func(it):
            lb = it[0]
            val = it[1]
            form = [l.form for l in self.cases[index][1:] if l.label == lb][0]
            form = form.split(',')
            formatted = []
            for elem in form:
                if elem == 'pct':
                    formatted.append(f'{val * 100}%')
                elif contains_any(elem, COMPARISON_OPERATORS):
                    formatted.append(eval(elem))
                elif elem == 'int':
                    formatted.append(val)
            return lb, ', '.join([str(v) for v in formatted])

        month, year = get_date(self.file_name)
        logging.info('-----------------------')
        logging.info(f'Data for {month}/{year}')
        for index in range(len(self.cases)):
            data = self.collect_data(index, month, year)
            data.sort(key=sort_func)
            data = list(map(map_func, data))
            for label, d in data:
                logging.info(f'{label}: {d}')


def main():
    case1 = [TEST_SHEET[0]]
    for tup in TEST_SHEET[1:]:
        case1.append(DataPoint(*tup))
    case2 = [VOC_SHEET[0]]
    for tup in VOC_SHEET[1:]:
        case2.append(DataPoint(*tup))
    c = DataLogger(file_to_open, [case1, case2])
    c.log_data()


if __name__ == '__main__':
    main()


# fix the data in the march spreadsheet to match the correct format
def fix_data():
    filename = 'expedia_report_monthly_march_2018.xlsx'
    wb = xl.open(filename)
    sheet = wb['VOC Rolling MoM']
    sheet['B1'] = dt.date(2018, 3, 1)
    sheet['C1'] = dt.date(2018, 2, 1)
    sheet['D1'] = dt.date(2018, 1, 1)
    wb.save(filename)


# fix_data()
